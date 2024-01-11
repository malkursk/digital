from django.http import HttpResponse
import openpyxl
from my.models import *



def export_to_excel(request, v):
    model = None
    fields = None
    match v:
        case "sport":
            model = Sport
            fields = ['id','name']
        case "game":
            model = Game
        case "news":
            model = News        
        case "person":
            model = Person
            fields = ['id','first_name','last_name','born','address']
        case "winner":
            model = Winner        
        case "user":
            model = get_user_model()
            fields = ['id','username','first_name','last_name','email','is_staff','is_active','groups']
            
    if not model:
        return HttpResponse('Экспорт невозможен: указана неверная модель')    

    wb = openpyxl.Workbook()     
    ws  = wb.active     
    ws.title = v
    col = row = 1
    if not fields:
        fields = list([field.name for field in model._meta.get_fields()]);
    print('param: ' + v)
    print(fields)
    for cell_header in fields:
        ws.cell(1, col, cell_header)    
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = '25'
        cell = ws.cell(1, col)        
        cell.fill = openpyxl.styles.PatternFill(start_color='DEEC91', end_color='DEEC91', fill_type="solid")  
        col+=1

    row += 1
    df = model.objects.values_list(*fields)
    for row_data in df:
        col = 1
        for cell_data in row_data:
            ws.cell(row, col, str(cell_data))            
            col += 1
        row += 1
    response = HttpResponse(content_type='application/vnd.ms-excel')        
    wb.save(response)
    return response