from django.shortcuts import redirect, render, HttpResponse
from openpyxl import load_workbook

from my.models import Person

def import_person_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['file']
        wb = load_workbook(excel_file)
        ws = wb.active
        if not ws.title == 'person':
            return HttpResponse("Импорт невозможен: лист excel должен называться person") 
        cnt_updated = 0
        cnt_created = 0
        cnt_skipped = 0
        cnt_all = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            cnt_all += 1
            id = row[0]
            first_name = row[1]
            last_name = row[2]
            born = row[3]
            address = row[4]            
            if first_name and last_name and born and address:
                if id.isnumeric() and (Person.objects.filter(id=id)):
                    if Person.objects.filter(id=id, first_name=first_name, last_name=last_name, born=born, address=address):
                        cnt_skipped += 1
                    else:
                        Person.objects.filter(id=id).update(first_name=first_name, last_name=last_name, born=born, address=address)
                        cnt_updated += 1
                else:
                    Person.objects.create(first_name=first_name, last_name=last_name, born=born, address=address)
                    cnt_created += 1

        return HttpResponse(f"<p>Импорт завершен, записей:  {cnt_all}, из них: </p><li>пропущено (полностью совпадают): {cnt_skipped}<li>обновлено (внесены уточнения): {cnt_updated}<li>создано новых: {cnt_created}") 
    return render(request, 'my/import_form.html')